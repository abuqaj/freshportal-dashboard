"""Generate Excel report from verification results."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from verifier import VerificationResult


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def generate_report(results: list[VerificationResult], output_path: Path) -> None:
    errors = [r for r in results if r.status in ("ERROR", "WARNING")]
    all_rows = []
    err_rows = []

    for r in results:
        official_name = r.vbn_info.official_name if r.vbn_info else ""
        vbn_group = r.vbn_info.product_group if r.vbn_info else ""

        all_rows.append({
            "ID produktu": r.product.product_id,
            "Kod produktu": r.product.short_name,
            "Nazwa": r.product.name,
            "Aktualny VBN": r.product.vbn_number,
            "Nazwa wg VBN": official_name,
            "Grupa VBN": vbn_group,
            "Status": r.status,
            "Powód": r.reason,
            "Proponowany VBN": r.proposed_vbn,
        })

        if r.status in ("ERROR", "WARNING"):
            err_rows.append({
                "ID produktu": r.product.product_id,
                "Kod produktu": r.product.short_name,
                "Nazwa": r.product.name,
                "Aktualny VBN": r.product.vbn_number,
                "Nazwa wg VBN": official_name,
                "Proponowany VBN": r.proposed_vbn,
                "Powód": r.reason,
            })

    df_errors = pd.DataFrame(err_rows)
    df_all = pd.DataFrame(all_rows)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_errors.to_excel(writer, sheet_name="Błędy", index=False)
        df_all.to_excel(writer, sheet_name="Wszystkie", index=False)

        wb = writer.book

        # Style "Błędy" sheet
        ws_err = wb["Błędy"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="C0392B")
        for cell in ws_err[1]:
            cell.font = header_font
            cell.fill = header_fill
        _auto_width(ws_err)

        # Style "Wszystkie" sheet
        ws_all = wb["Wszystkie"]
        header_fill_all = PatternFill(fill_type="solid", fgColor="2C3E50")
        for cell in ws_all[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill_all

        # Colour status column rows
        status_col = None
        for i, cell in enumerate(ws_all[1], start=1):
            if cell.value == "Status":
                status_col = i
                break

        if status_col:
            ok_fill = PatternFill(fill_type="solid", fgColor="D5E8D4")
            err_fill = PatternFill(fill_type="solid", fgColor="F8CECC")
            warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
                cell = row[status_col - 1]
                if cell.value == "OK":
                    cell.fill = ok_fill
                elif cell.value == "ERROR":
                    cell.fill = err_fill
                elif cell.value == "WARNING":
                    cell.fill = warn_fill

        _auto_width(ws_all)

    print(f"\nReport saved: {output_path}")
    print(f"  Total products: {len(results)}")
    print(f"  Errors/Warnings: {len(errors)}")
    print(f"  OK: {len(results) - len(errors)}")
